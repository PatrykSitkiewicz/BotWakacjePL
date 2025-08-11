

# Weryfikacja niezbędnych bibliotek w systemie
# ------------------------------------------------------------
import subprocess
import sys

# BotWakacjePLv3.0.py

# scraper.py
import os
import re
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


def scrape_hotels_and_ratings(wakacje_link_template: str, num_pages: int, chromedriver_path: str, excel_output_path: str):
    """
    Główna funkcja: zbiera dane z wakacje.pl i oceny z Google, zapisuje je w pliku Excel.
    """
    chrome_options = Options()
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    service = Service(chromedriver_path)
    driver = webdriver.Chrome(service=service, options=chrome_options)

    hotels_data = {}

    try:
        for X in range(1, num_pages + 1):
            url = wakacje_link_template.replace("/?str-2,", f"/?str-{X},")
            driver.get(url)

            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "div h4"))
                )
            except:
                print(f"Nie udało się załadować strony {X}")
                continue

            for _ in range(5):
                driver.execute_script("window.scrollBy(0, 1000);")
                time.sleep(1)

            offers = driver.find_elements(By.CSS_SELECTOR, 'a[data-test-offer-id]')

            for offer in offers:
                try:
                    # Nazwa hotelu
                    name_el = offer.find_element(By.CSS_SELECTOR, "[data-testid='offer-listing-name']")
                    hotel_name = name_el.text.strip()

                    # Lokalizacja hotelu
                    try:
                        location_el = offer.find_element(By.CSS_SELECTOR, "[data-testid='offer-listing-geo']")
                        location = location_el.text.strip().replace(" / ", ", ")
                    except:
                        location = None

                    # Cena hotelu
                    try:
                        price_element = offer.find_element(By.CSS_SELECTOR, 'h4[data-testid="txt"]')
                        price_text = price_element.text
                    except Exception:
                        price_text = ""
                    price_match = re.search(r"(\d[\d\s]*\d)\s*zł", price_text)
                    price = int(price_match.group(1).replace(" ", "")) if price_match else None

                    if hotel_name and price:
                        if hotel_name in hotels_data:
                            current = hotels_data[hotel_name]["price"]
                            hotels_data[hotel_name]["price"] = min(current, price)
                        else:
                            hotels_data[hotel_name] = {"price": price, "location": location}

                        print(f"Zapisano: {hotel_name} | {location} | {price} PLN")

                except Exception as e:
                    print(f"❌ Błąd przy analizie oferty: {e}")
                    continue



    finally:
        print("→ hotels_data:")
        for hotel, info in hotels_data.items():
            print("  ", hotel, "==>", info)
        _save_to_excel(hotels_data, excel_output_path)
        try:
            scrap_google_ratings(excel_output_path, driver)
        except Exception as e:
            print(f"❌ Błąd podczas scrapowania ocen Google: {e}")

        driver.quit()


def _save_to_excel(hotels_data: dict, excel_file: str):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Dane hotelowe"
    sheet.append(["Hotel", "Lokalizacja", "Cena"])

    for hotel_name, info in hotels_data.items():
        if isinstance(info, dict):
            location = info.get("location", "")
            price = info.get("price", "")
        else:
            location = ""
            price = info
        sheet.append([hotel_name, location, price])

    workbook.save(excel_file)
    print(f"Dane zapisano do pliku: {excel_file}")


def scrap_google_ratings(excel_file_path, driver):
    driver.get("https://www.google.com/search?q=wakacje+hotele")
    time.sleep(60)  # czas na rozwiązanie Captcha

    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]

    if "Ocena Google" not in headers:
        sheet.cell(row=1, column=4, value="Ocena Google")
        headers.append("Ocena Google")

    google_col_index = headers.index("Ocena Google") + 1

    for row_num in range(2, sheet.max_row + 1):
        hotel_name = sheet.cell(row=row_num, column=1).value
        location = sheet.cell(row=row_num, column=2).value  # ⬅️ dodaj to

        if not hotel_name:
            continue

        # Bezpiecznie obsłuż brak lokalizacji
        location_str = location.replace(" / ", ", ").replace(" ", "+") if location else ""

        query = hotel_name.replace(" ", "+") + ",+hotel+" + location_str
        url = f"https://www.google.com/search?q={query}"
        driver.get(url)
        time.sleep(2)

        try:
            rating_element = driver.find_element(By.CSS_SELECTOR, ".Aq14fc")
            rating_text = rating_element.text.strip().replace(",", ".")
            rating_value = float(rating_text)
        except Exception:
            rating_value = None  # brak oceny

        sheet.cell(row=row_num, column=google_col_index, value=rating_value)
        print(f"[Ocena Google] {hotel_name} -> {rating_value if rating_value is not None else 'Brak'}")

    workbook.save(excel_file_path)
    print("Oceny Google zapisane.")
