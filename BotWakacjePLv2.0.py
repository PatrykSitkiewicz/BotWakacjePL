#--------------------------------
# Name:        BotWakacjePL 2.0
# Author:   Patryk Sitkiewicz
# Created     04 December 2024
# Last update: 17 February 2025
# Python Version:   3.13
#--------------------------------
### Przed użyciem przeczytaj instrukcję w README.md
# Bot przegląda wskazaną stronę wyszukiwania hoteli na portalu Wakacje.pl
# Zapisuje nazwy hoteli wraz z ich aktualnie najniższymi cenami do pliku excel
# następnie w Google sprawdza oceny hoteli i zapisuje je w kolejnej kolumnie
# uruchomiony ponownie aktualizuje ceny, a plik wynikowy umożliwia śledzenie zmian cen
# ------------------------------------------------------------

# Weryfikacja niezbędnych bibliotek w systemie
# ------------------------------------------------------------
import subprocess
import sys

# Lista wymaganych bibliotek do poprawego działania bota
required_libraries = ["selenium", "openpyxl"]

# Sprawdzanie, czy wymagane biblioteki są już zainstalowane
def install_and_import(package):
    try:
        __import__(package)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])

# Instalowanie brakujących bibliotek
for library in required_libraries:
    install_and_import(library)

# ------------------------------------------------------------
# Przygotowanie środowiska pracy
# ------------------------------------------------------------
import os
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import time
from datetime import datetime

# Konfiguracja opcji przeglądarki
chrome_options = Options()
#chrome_options.add_argument("--headless")  # Opcja: usunięcie tej linii sprawi, że okno przeglądarki Chrome będzie widoczne w trakcie pracy bota
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--no-sandbox")

# Ścieżka do ChromeDrivera
webdriver_path = r"C:\chromedriver-win64\chromedriver.exe" #UWAGA! w cudzysłowie musi znajdować się poprawna ścieżka do pliku "chromedriver.exe" zainstalowanego uprzednio

# Ścieżka do pliku Excela
excel_file = "hotels_data.xlsx" # tu opcjonalnie można zmienić nazwę oraz ścieżkę pliku wynikowego w formacie MS Excel

# Inicjalizacja przeglądarki
service = Service(webdriver_path)
driver = webdriver.Chrome(service=service, options=chrome_options)

# Słownik do przechowywania nowych danych
hotels_data = {}

# ------------------------------------------------------------
# Bot automatycznie przegląda za nas stronę Wakacje.pl i pobiera nazwy hoteli oraz ich ceny
# ------------------------------------------------------------
# # UWAGA!!! WAŻNE!!!
# poniżej znajduje adres strony Wakacje.pl, która ma być przeglądana przez bota
# zmień ją wklejając do cudzysłowu url = f"adres strony"; ta strona będzie przeglądana przez bota
# pamiętaj, że numer strony należy zamienić na wyrażenie {X}, jak w domyślnym adresie strony wydocznym poniżej
# wszystko po to, gdyż wyniki Wakacje.pl wyświetlane są na wielu podstronach - bot odczyta wszystkie dzięki pętli for
# domyślnie odczytuje wyniki na pierwszych 5 stronach, jeśli chcesz to zmienić wstaw inną wartość zamiast "6" w wyrażeniu for X in range(1, 6):
try:
    for X in range(1, 15):
        url = f"https://www.wakacje.pl/wczasy/?str-{X},od-2025-06-10,do-2025-06-29,samolotem,all-inclusive,z-gdanska,2dorosle-1dziecko-20220518,tanio" # tu zmień link!!
        print(f"Przetwarzanie strony: {url}")

        # Otwórz wskazaną stronę w portalu Wakacje.pl
        driver.get(url)

        # Załadowanie strony
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div h4"))
            )
        except:
            print(f"Nie udało się załadować elementów na stronie {X}")
            continue

        # Bot przewija stronę, by wczytała się w całości
        for _ in range(5):  # Liczba przewinięć strony
            driver.execute_script("window.scrollBy(0, 1000);")
            time.sleep(1)

        # Pobierz dane z elementów "div h4", w których na Wakacje.pl znajdują się nazwy hoteli oraz ich ceny
        raw_data = driver.find_elements(By.CSS_SELECTOR, "div h4")

        # Przetwarzanie danych
        hotel_name = None
        for element in raw_data:
            text = element.text.strip()

            # Rozbij tekst na linie, aby usunąć pobrane wyrażenie "od" i oddzielić cenę od reszty wyników
            lines = text.split("\n")
            for line in lines:
                line = line.strip()
                if line == "od":
                    continue  # Pomijamy linie "od"

                # Sprawdzamy, czy pobrane wyrażenie jest ceną
                price_match = re.match(r"^(\d[\d\s]*\d)\s*zł$", line)
                if price_match:
                    price = int(price_match.group(1).replace(" ", ""))  # Usuwamy spacje z liczby
                    if hotel_name:
                        # Zapisz hotel i jego najniższą cenę w słowniku
                        if hotel_name in hotels_data:
                            hotels_data[hotel_name] = min(hotels_data[hotel_name], price)
                        else:
                            hotels_data[hotel_name] = price
                        print(f"Zapisano: Hotel - {hotel_name}, Cena - {price} PLN")
                        hotel_name = None  # Reset nazwy hotelu
                else:
                    # Jeśli nie jest ceną, traktujemy wyrażenie jako nazwę hotelu
                    hotel_name = line
# ------------------------------------------------------------
# Bot zapisuje przetworzone nazwy hoteli wraz z ich najniższymi cenami do pliku MS Excel
# ------------------------------------------------------------
finally:
    # Obsługa pliku Excel
    if os.path.exists(excel_file):
        # Otwórz istniejący plik
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # Sprawdź, czy istnieje kolumna "Cena dzisiaj"
        headers = [cell.value for cell in sheet[1]]
        if "Cena dzisiaj" not in headers:
            sheet.cell(row=1, column=len(headers) + 1, value="Cena dzisiaj")
            headers.append("Cena dzisiaj")

        # Dodaj nowe ceny do istniejących hoteli
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            hotel_name = row[0].value  # Nazwa hotelu
            if hotel_name in hotels_data:
                # Znajdź kolumnę dla "Cena dzisiaj" i zaktualizuj cenę, jeśli jest niższa
                col_index = headers.index("Cena dzisiaj") + 1
                current_price = row[col_index - 1].value
                new_price = hotels_data[hotel_name]
                if current_price is None or new_price < current_price:
                    row[col_index - 1].value = new_price
                del hotels_data[hotel_name]  # Usuń zapisany hotel z nowego słownika

        # Dodaj nowe hotele (których nie było w istniejącym pliku)
        for hotel_name, price in hotels_data.items():
            new_row = [hotel_name] + [None] * (len(headers) - 2) + [price]
            sheet.append(new_row)

    else:
        # Stwórz nowy plik, jeśli nie istnieje
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Dane hotelowe"
        sheet.append(["Hotel", "Cena", "Cena dzisiaj"])
        for hotel_name, price in hotels_data.items():
            sheet.append([hotel_name, price, price])

    # Zapisz zmiany w pliku Excel
    workbook.save(excel_file)
    print(f"Dane zapisano w pliku Excel: {excel_file}")

    # Zamknij przeglądarkę
    driver.quit()


# ---------------------------------------------------------------------
# NOWA FUNKCJA - dodana w wersji 2.0 - niezależnie SPRAWDZA OCENY HOTELI W GOOGLE
# działa niezależnie od pierwszej funkcji bota
# ---------------------------------------------------------------------

def scrap_google_ratings(excel_file_path):
    """
    Funkcja odczytuje z pliku Excel nazwy hoteli,
    wyszukuje ich oceny w Google (klasa .Aq14fc) i dopisuje wyniki
    do nowej kolumny 'Ocena Google' (4-ta kolumna).
    """

    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    import time
    import openpyxl

    # --- Konfiguracja nowej przeglądarki (żeby nie modyfikować starej) ---
    chrome_options = Options()
    #chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")

    # Otwiera przeglądarkę i przechodzi do Google, by sprawdzić, czy jest Capcha, jeśli tak, masz 60 sekund na jej rozwiązanie ---
    new_service = Service(webdriver_path)  # Używamy tej samej ścieżki do chromedrivera
    new_driver = webdriver.Chrome(service=new_service, options=chrome_options)
    new_driver.get(f"https://www.google.com/search?q=test")
    time.sleep(60)

    # Otwieramy plik Excel
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active  # Zakładamy, że chcemy pisać w tym samym arkuszu

    # Sprawdzamy, czy mamy już kolumnę "Ocena Google"
    headers = [cell.value for cell in sheet[1]]
    if "Ocena Google" not in headers:
        # Dodajemy w czwartej kolumnie
        # (A=Hotel, B=Cena, C=Cena dzisiaj, D=Ocena Google)
        sheet.cell(row=1, column=4, value="Ocena Google")
        headers.append("Ocena Google")

    # Indeks kolumny "Ocena Google"
    google_col_index = headers.index("Ocena Google") + 1  # +1 bo index w Python, a numer kolumny w Excelu

    # Przechodzimy po wierszach, czytamy nazwy hoteli
    max_row = sheet.max_row
    for row_num in range(2, max_row + 1):
        hotel_name = sheet.cell(row=row_num, column=1).value  # 1 = kolumna A (Hotel)
        if not hotel_name:
            continue  # Pusty wiersz

        # Przygotowanie zapytania do Google
        query = hotel_name.replace(" ", "+") + "+hotel" #  zmień na właściwe! dla lepszych wyników obok +hotel dodaj + oraz zaraz ponim nazwę kraju, w którym hotele sprawdzasz
        search_url = f"https://www.google.com/search?q={query}"

        # Bot wchodzi na stronę z wynikami Google
        new_driver.get(search_url)
        time.sleep(2)  # krótka pauza, żeby Google zdążyło załadować wyniki

        # Znalezienie div .Aq14fc
        try:
            rating_element = new_driver.find_element(By.CSS_SELECTOR, ".Aq14fc")
            rating_text = rating_element.text.strip()
        except Exception:
            rating_text = "Brak oceny"

        # Zapisz ocenę do Excela
        sheet.cell(row=row_num, column=google_col_index, value=rating_text)
        print(f"[Ocena Google] {hotel_name} -> {rating_text}")

    # Zapisz zmiany i zamknij przeglądarkę
    workbook.save(excel_file_path)
    new_driver.quit()

    print(f"\nWszystkie oceny Google zostały dopisane do pliku {excel_file_path} w kolumnie 'Ocena Google'.")


# ---------------------------------------------------------------------
# Wywołanie nowej funkcji PO wykonaniu BotWakacjePL
# ---------------------------------------------------------------------
if __name__ == "__main__":
    # Po zakończeniu pracy BotWakacjePL wywołujemy funkcję scrap_google_ratings.
    scrap_google_ratings(excel_file)